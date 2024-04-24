from django.shortcuts import render, redirect
from django.http import HttpResponse
import requests
import json
from .models import *
from core.models import *
from django.contrib import messages
from django.urls import reverse
from openpyxl import load_workbook


# Create your views here.


def exams_dashboard(request):
    return render(request, 'exams/dashboard.html')


def upload_result(request):
    if request.method == 'POST':
        session_id = request.POST['session']
        semester_id = request.POST['semester']
        course_id = request.POST['course']
        excel_file = request.FILES['file_upload']
        wb = load_workbook(excel_file)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            student, ca, exams = row
            exam_result = ExamResult(
                student = student,
                session = Session.objects.get(id=session_id),
                semester = Semester.objects.get(id=semester_id),
                course = Course.objects.get(id=course_id),
                ca = ca,
                exams = exams
            )
            exam_result.save()
        messages.success(request, 'Result Uploaded Successfully')
        return redirect('upload-result')
    
    context = {
        'sessions': Session.objects.all(),
        'semesters': Semester.objects.all(),
        'courses': Course.objects.all(),
    }

    return render(request, 'exams/upload_result.html', context)


def update_result(request):
    if request.method == 'POST':
        session_id = request.POST['session']
        semester_id = request.POST['semester']
        course_id = request.POST['course']
        excel_file = request.FILES['file_upload']
        wb = load_workbook(excel_file)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            student, ca, exams = row
            try:
                exam_result = ExamResult.objects.get(
                    student = student,
                    session = Session.objects.get(id=session_id),
                    semester = Semester.objects.get(id=semester_id),
                    course = Course.objects.get(id=course_id),
                )
                exam_result.ca = ca
                exam_result.exams = exams
                exam_result.save()
                mm = "success"
            except ExamResult.DoesNotExist:
                mm = "error"
                pass
        if mm == "success":
            messages.success(request, 'Result Updated Successfully')
        else:
            messages.error(request, 'Some of the Result May not Exist')
        return redirect('update-result')
    
    context = {
        'sessions': Session.objects.all(),
        'semesters': Semester.objects.all(),
        'courses': Course.objects.all(),
    }

    return render(request, 'exams/update_result.html', context)


